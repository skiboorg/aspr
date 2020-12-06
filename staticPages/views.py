from django.shortcuts import render,get_object_or_404,HttpResponse
from .models import *
from openpyxl import load_workbook
# import urllib
# from urllib.parse import urlparse
# import requests
# from django.core.files.base import ContentFile
from shop.models import *

def avtomatizaciya(request):
    all_cats = Category.objects.all()
    page_title = 'АСПР Групп | Автоматизация'
    page_description = 'АСПР Групп | Санкт-Петербург | Промышленная автоматизация'
    page_content = HomePage.objects.all().first()
    content = AvtomatizaciyaPage.objects.all().first()
    avt = True
    canonical_url = '/avtomatizaciya'
    return render(request, 'pages/page.html', locals())

def proizvodstvo_item(request,name_slug):
    all_cats = Category.objects.all()
    page_title = 'АСПР Групп | Производство'
    page_description = 'АСПР Групп | Санкт-Петербург | Промышленная автоматизация'
    content = get_object_or_404(Manufactory,name_slug=name_slug)
    page_content = HomePage.objects.all().first()
    pr = True
    canonical_url = f'/proizvodstvo/{name_slug}'
    return render(request, 'pages/page.html', locals())
def proizvodstvo(request):
    showitems = True
    all_cats = Category.objects.all()
    page_title = 'АСПР Групп | Производство'
    page_description = 'АСПР Групп | Санкт-Петербург | Промышленная автоматизация'
    page_content = HomePage.objects.all().first()
    content = ProizvodsvoPage.objects.all().first()
    pr = True
    canonical_url = '/proizvodstvo'
    return render(request, 'pages/page.html', locals())

def projects(request):
    prj = True
    all_cats = Category.objects.all()
    page_title = 'АСПР Групп | Проекты'
    page_description = 'АСПР Групп | Санкт-Петербург | Промышленная автоматизация'
    all_projects = Project.objects.all()
    canonical_url = '/projects'
    return render(request, 'pages/projects.html', locals())


def napravleniya(request):
    all_cats = Category.objects.all()
    nap = True
    page_title = 'АСПР Групп | Направления'
    page_description = 'АСПР Групп | Санкт-Петербург | Промышленная автоматизация'
    content = WorkTypePage.objects.all().first()
    canonical_url = '/napravleniya'
    return render(request, 'pages/worktype.html', locals())

# def create_item(request):
#
#     url = '/home/xxx/'
#     wb = load_workbook(filename=f'{url}runn.xlsx')
#     sheet = wb.active
#     max_row = sheet.max_row
#     max_column = sheet.max_column
#
#
#
#
#     for i in range(3, max_row + 1):
#         article = sheet.cell(row=i, column=1).value
#
#         if not article:
#             break
#         inputs = sheet.cell(row=i, column=2).value
#         nominal_tok = sheet.cell(row=i, column=3).value
#         is_avr = sheet.cell(row=i, column=4).value
#         manufactor = sheet.cell(row=i, column=5).value
#         defence = sheet.cell(row=i, column=6).value
#         price = sheet.cell(row=i, column=7).value
#         setup_type = sheet.cell(row=i, column=8).value
#
#         warranty = sheet.cell(row=i, column=9).value
#         life_time = sheet.cell(row=i, column=10).value
#         work_temperature = sheet.cell(row=i, column=11).value
#         kabel_input = sheet.cell(row=i, column=12).value
#
#
#         print(price)
#
#
#
#         try:
#             manufactor = Manufactor.objects.get(name=manufactor)
#         except:
#             manufactor = Manufactor.objects.create(name=manufactor)
#
#         full_d="""
#         """
#
#         new_item = Item.objects.create(
#             name='РУНН',
#             subcategory_id=8,
#
#             price=int(price),
#             article=article,
#             inputs=inputs,
#             nominal_tok=nominal_tok,
#             is_avr = True if is_avr == 'Да' else False,
#             # is_counters = True if is_counters == 'Да' else False,
#             manufactor=manufactor,
#             # power=power,
#             # tok_str=tok_str,
#             # number_faz=number_faz,
#             # napryag=napryag
#              defence=defence,
#             # number_zadvigek=number_zadvigek,
#             # engine_tok=engine_tok,
#              setup_type=setup_type,
#              #size_shelter=size_shelter,
#              warranty=warranty,
#              life_time=life_time,
#
#              work_temperature=work_temperature,
#              kabel_input=kabel_input,
#
#              # full_description=full_d,
#             image='item/runn.jpg'
#             )
#
#
#
#
#     return render(request, 'page/about.html', locals())
def about(request):
    ab = True
    all_cats = Category.objects.all()
    page_title = 'АСПР Групп | О нас'
    page_description = 'АСПР Групп | Санкт-Петербург | Промышленная автоматизация'
    content = AboutPage.objects.all().first()
    canonical_url = '/about'
    return render(request, 'pages/about.html', locals())

def index(request):
    ind = True
    all_cats = Category.objects.all()
    page_title = 'АСПР Групп | Санкт-Петербург | Промышленная автоматизация'
    page_description = 'АСПР Групп | Санкт-Петербург | Промышленная автоматизация'
    page_content = HomePage.objects.all().first()
    canonical_url = ''
    return render(request, 'pages/index.html', locals())

def customhandler404(request, exception, template_name='404.html'):
    all_cats = Category.objects.all()
    is404 = True
    pageTitle = '404 - Такой страницы не существует'


    return render(request, 'pages/404.html', locals(),None,status=404)

def robots(request):
    robotsTxt = f"User-agent: *\nDisallow: /admin/\nHost: https://www.aspr.tech\nSitemap: https://www.aspr.tech/sitemap.xml"
    return HttpResponse(robotsTxt, content_type="text/plain")
